# ============================================================
# EXTRACTOR JIRA - STREAMLIT
# Equipo: Valeria, pcateriano, Agustin, Marcelo, Moisés,
#         rvega, Cecilia, Junior
# Credenciales via st.secrets
# ============================================================

import io
import time
import requests
import pandas as pd
import streamlit as st
from requests.auth import HTTPBasicAuth
from datetime import datetime

# ============================================================
# CONFIGURACIÓN ESTÁTICA
# ============================================================

PROJECT_KEYS = ['PM']

ASSIGNEES = [
    "valeria vergaray",
    "pcateriano",
    "Agustin Gutierrez",
    "Marcelo Reyes",
    "Moisés Ayala",
    "rvega",
    "Cecilia Puente",
    "Junior Suasnabar",
]

ALL_STATUSES = [
    "Por hacer"
    "BACKLOG UX",
    "EN CURSO DE UX",
    "BACKLOG SOFTWARE | COE",
    "BACKLOG SOFTWARE",
    "EN CURSO DE SOFTWARE | COE",
    "ATENDIDO",
    "A LA ESPERA DE LEGAL",
    "A LA ESPERA DE BA/DATA",
    "A LA ESPERA DE MARKETING",
    "A LA ESPERA DE TERCEROS",
    "DESECHADO"
]

# ============================================================
# CREDENCIALES DESDE st.secrets
# ============================================================

JIRA_URL  = st.secrets["JIRA_URL"]
USERNAME  = st.secrets["JIRA_USER"]
API_TOKEN = st.secrets["JIRA_TOKEN"]

# ============================================================
# CLASE: EXTRACTOR DE ISSUES
# ============================================================

class JiraExtractor:
    def __init__(self):
        self.auth    = HTTPBasicAuth(USERNAME, API_TOKEN)
        self.headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}
        self.base    = JIRA_URL.rstrip('/')

    def _post(self, url, payload):
        r = requests.post(url, json=payload, auth=self.auth, headers=self.headers)
        r.raise_for_status()
        return r.json()

    def _get(self, url, params=None):
        r = requests.get(url, params=params, auth=self.auth, headers=self.headers)
        r.raise_for_status()
        return r.json()

    def test_connection(self):
        try:
            info = self._get(f"{self.base}/rest/api/2/myself")
            return True, info.get('displayName', '')
        except Exception as e:
            return False, str(e)

    def fetch_issues(self, project_key, date_from_str, date_to_str, log_fn=None):
        url = f"{self.base}/rest/api/3/search/jql"
        assignees_jql = ', '.join(f'"{a}"' for a in ASSIGNEES)
        jql = (
            f'project = "{project_key}" '
            f'AND assignee in ({assignees_jql}) '
            f'AND created >= "{date_from_str}" '
            f'AND created <= "{date_to_str}" '
            f'ORDER BY created DESC'
        )
        all_issues = []
        next_token = None
        page = 0

        while True:
            payload = {
                "jql": jql,
                "maxResults": 50,
                "fields": [
                    "key", "summary", "status", "assignee",
                    "created", "updated", "issuetype",
                    "resolution", "resolutiondate", "description", "labels",
                    "customfield_12066", "customfield_12067", "customfield_12166"
                ]
            }
            if next_token:
                payload["nextPageToken"] = next_token

            data   = self._post(url, payload)
            issues = data.get('issues', [])
            if not issues:
                break

            page += 1
            for issue in issues:
                all_issues.append(self._parse_issue(issue, project_key))
            if log_fn:
                log_fn(f"Página {page}: {len(all_issues)} issues acumulados")

            next_token = data.get('nextPageToken')
            if not next_token:
                break
            time.sleep(0.3)

        return all_issues

    def _parse_issue(self, issue, project_key):
        f = issue.get('fields', {})

        def sg(obj, *keys):
            r = obj
            for k in keys:
                r = r.get(k) if isinstance(r, dict) else None
            return r or ''

        def fmt_dt(s):
            return (s[:10] + ' ' + s[11:19]) if s and 'T' in s else (s or '')

        def fmt_d(s):
            return s[:10] if s else ''

        def clean_desc(d):
            if not d: return ''
            if isinstance(d, dict):
                parts = []
                for block in d.get('content', []):
                    for sub in block.get('content', []):
                        if sub.get('text'): parts.append(sub['text'])
                return ' '.join(parts)[:400]
            return str(d)[:400]

        def parse_arr(fd):
            if not fd: return ''
            if isinstance(fd, list):
                return ', '.join(
                    (i.get('value', str(i)) if isinstance(i, dict) else str(i))
                    for i in fd
                )
            return str(fd)

        def parse_single(fd):
            if not fd: return ''
            return fd.get('value', '') if isinstance(fd, dict) else str(fd)

        return {
            'proyecto_codigo':        project_key,
            'issue_key':              issue.get('key', ''),
            'url_ticket':             f"https://prestamype.atlassian.net/browse/{issue.get('key', '')}",
            'issue_id':               issue.get('id', ''),
            'summary':                f.get('summary', ''),
            'description':            clean_desc(f.get('description')),
            'status':                 sg(f, 'status', 'name') or 'Sin estado',
            'issue_type':             sg(f, 'issuetype', 'name') or 'Sin tipo',
            'assignee':               sg(f, 'assignee', 'displayName') or 'Sin asignar',
            'created_date':           fmt_d(f.get('created', '')),
            'created_datetime':       fmt_dt(f.get('created', '')),
            'updated_date':           fmt_d(f.get('updated', '')),
            'updated_datetime':       fmt_dt(f.get('updated', '')),
            'resolution':             sg(f, 'resolution', 'name') or 'Sin resolver',
            'resolution_date':        fmt_d(f.get('resolutiondate', '')),
            'resolution_datetime':    fmt_dt(f.get('resolutiondate', '')),
            'labels':                 ', '.join(f.get('labels', [])),
            'categoria_AQN':          parse_arr(f.get('customfield_12066')),
            'PMBOK':                  parse_single(f.get('customfield_12067')),
            'informacion_completada': parse_single(f.get('customfield_12166')),
        }


# # ============================================================
# # CLASE: EXTRACTOR DE CHANGELOG  (comentado — no requerido)
# # ============================================================
#
# class ChangelogExtractor:
#     def __init__(self):
#         self.auth    = HTTPBasicAuth(USERNAME, API_TOKEN)
#         self.headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}
#         self.base    = JIRA_URL.rstrip('/')
#
#     def fetch_changelog(self, issue_key):
#         url    = f"{self.base}/rest/api/3/issue/{issue_key}/changelog"
#         result = []
#         start  = 0
#         while True:
#             try:
#                 r = requests.get(
#                     url, params={'startAt': start, 'maxResults': 100},
#                     auth=self.auth, headers=self.headers
#                 )
#                 if r.status_code != 200:
#                     break
#                 data   = r.json()
#                 values = data.get('values', [])
#                 if not values:
#                     break
#                 for entry in values:
#                     created = entry.get('created', '')
#                     author  = entry.get('author', {}).get('displayName', 'Unknown')
#                     for item in entry.get('items', []):
#                         if item.get('field', '').lower() == 'status':
#                             result.append({
#                                 'issue_key':   issue_key,
#                                 'from_status': item.get('fromString', ''),
#                                 'to_status':   item.get('toString', ''),
#                                 'change_dt':   created[:19].replace('T', ' ') if len(created) >= 19 else created,
#                                 'changed_by':  author,
#                             })
#                 if data.get('isLast', True):
#                     break
#                 start += 100
#             except Exception:
#                 break
#         return result
#
#     def fetch_all(self, issue_keys, progress_bar=None, log_fn=None):
#         all_changes = []
#         total = len(issue_keys)
#         for i, key in enumerate(issue_keys, 1):
#             all_changes.extend(self.fetch_changelog(key))
#             if progress_bar:
#                 progress_bar.progress(i / total, text=f"Changelog {i}/{total} issues")
#             if log_fn and (i % 20 == 0 or i == total):
#                 log_fn(f"Changelog {i}/{total}")
#             time.sleep(0.12)
#         return all_changes


# ============================================================
# FILTRO DE ISSUES (solo por estado)
# ============================================================

def apply_filters(issues, selected_statuses):
    if not selected_statuses:
        return issues
    filtered = []
    for issue in issues:
        if issue['status'].upper().strip() in [s.upper().strip() for s in selected_statuses]:
            filtered.append(issue)
    return filtered


# ============================================================
# GENERADOR DE EXCEL
# ============================================================

def build_excel_bytes(issues):
    from openpyxl.styles import PatternFill, Font, Alignment

    df = pd.DataFrame(issues)

    col_order = [
        'proyecto_codigo', 'issue_key', 'url_ticket', 'issue_id', 'summary',
        'status', 'issue_type', 'assignee',
        'created_date', 'created_datetime',
        'updated_date', 'updated_datetime',
        'resolution', 'resolution_date', 'resolution_datetime',
        'labels', 'categoria_AQN', 'PMBOK', 'informacion_completada',
        'description',
    ]
    col_order = [c for c in col_order if c in df.columns]
    df = df[col_order]

    def style_header(ws, columns):
        for i, col in enumerate(columns, 1):
            cell           = ws.cell(row=1, column=i)
            cell.fill      = PatternFill("solid", fgColor="1a73e8")
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    col_w = {
        'proyecto_codigo': 16, 'issue_key': 14, 'issue_id': 12, 'url_ticket': 45, 'summary': 40,
        'status': 20, 'issue_type': 16, 'assignee': 22,
        'created_date': 14, 'created_datetime': 20,
        'updated_date': 14, 'updated_datetime': 20,
        'resolution': 16, 'resolution_date': 14, 'resolution_datetime': 20,
        'labels': 20, 'categoria_AQN': 22, 'PMBOK': 18,
        'informacion_completada': 24, 'description': 50,
    }

    # # Hoja Changelog (comentada — desactivada)
    # df_cl = pd.DataFrame(changelog_list)
    # df_cl.columns = ['Issue Key', 'De Estado', 'A Estado', 'Fecha y Hora', 'Modificado por']

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Issues')
        ws_i = writer.sheets['Issues']
        style_header(ws_i, df.columns)
        for i, col in enumerate(df.columns, 1):
            ws_i.column_dimensions[ws_i.cell(row=1, column=i).column_letter].width = col_w.get(col, 18)
        ws_i.freeze_panes = 'A2'

        # # Hoja Changelog (comentada — desactivada)
        # df_cl.to_excel(writer, index=False, sheet_name='Changelog')
        # ws_c = writer.sheets['Changelog']
        # ...

    output.seek(0)
    return output.read()


# ============================================================
# UI STREAMLIT
# ============================================================

st.set_page_config(page_title="Extractor JIRA", page_icon="📊", layout="wide")
st.title("Reporte AQN - Seguimiento Jira")

# ── Sidebar ─────────────────────────────────────────────────
with st.sidebar:
    st.header("Configuración")

    st.subheader("Creación del ticket")
    col_a, col_b = st.columns(2)
    with col_a:
        date_from = st.date_input("Desde", value=datetime(2025, 12, 1).date())
    with col_b:
        date_to   = st.date_input("Hasta", value=datetime(2025, 12, 31).date())

    st.subheader("Estado del ticket")
    selected_statuses = st.multiselect(
        "Estados (vacío = todos)",
        options=ALL_STATUSES,
        default=[]
    )

    st.subheader("Nombre del archivo")
    filename_input = st.text_input("Nombre (sin .xlsx)", value="REPORTE_AQN")

    run_btn = st.button("🚀 Ejecutar extracción", type="primary", use_container_width=True)

# ── Panel principal ──────────────────────────────────────────
if not run_btn:
    st.info("Configura los filtros en el panel izquierdo y presiona **Ejecutar extracción**.")
    st.stop()

if date_from > date_to:
    st.error("La fecha de inicio no puede ser mayor a la fecha fin.")
    st.stop()

# Conexión
extractor = JiraExtractor()
ok, display_name = extractor.test_connection()
if not ok:
    st.error(f"Error de conexión con JIRA: {display_name}")
    st.stop()
st.success(f"✅ Conectado como: {display_name}")

all_issues = []

# Paso 1: Issues
with st.status("Extrayendo issues...", expanded=True) as status_issues:
    for pk in PROJECT_KEYS:
        st.write(f"Proyecto: **{pk}**")
        issues = extractor.fetch_issues(
            pk,
            date_from.strftime('%Y-%m-%d'),
            date_to.strftime('%Y-%m-%d'),
            log_fn=lambda m: st.write(m)
        )
        all_issues.extend(issues)
    status_issues.update(
        label=f"✅ Issues extraídos: {len(all_issues)}",
        state="complete"
    )

if not all_issues:
    st.warning("No se encontraron issues en el rango de fechas seleccionado.")
    st.stop()

# # Paso 2: Changelog (comentado — desactivado)
# issue_keys = [i['issue_key'] for i in all_issues]
# ch_extractor = ChangelogExtractor()
# with st.status("Extrayendo changelog...", expanded=True) as status_cl:
#     prog_bar = st.progress(0, text="Iniciando...")
#     changelog = ch_extractor.fetch_all(issue_keys, progress_bar=prog_bar)
#     status_cl.update(
#         label=f"✅ Changelog extraído: {len(changelog)} cambios de estado",
#         state="complete"
#     )

# Paso 2: Filtros
with st.spinner("Aplicando filtros..."):
    filtered_issues = apply_filters(all_issues, selected_statuses)

if not filtered_issues:
    st.warning("Ningún issue pasó los filtros aplicados. Ajusta los criterios en el sidebar.")
    st.stop()

# Métrica
st.metric("Issues en reporte", len(filtered_issues))

# Paso 3: Generar Excel
with st.spinner("Generando archivo Excel..."):
    excel_bytes = build_excel_bytes(filtered_issues)

safe_name = "".join(c for c in filename_input.strip() if c.isalnum() or c in ('_', '-')).strip() or 'jira_report'
excel_filename = f"{safe_name}.xlsx"

st.download_button(
    label="⬇️ Descargar reporte Excel",
    data=excel_bytes,
    file_name=excel_filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)

# Tabla de resultados
st.subheader(f"📋 Resultados — {len(filtered_issues)} issues")

df_view = pd.DataFrame(filtered_issues)

col_order_view = [
    'proyecto_codigo', 'issue_key', 'url_ticket', 'issue_id', 'summary',
    'status', 'issue_type', 'assignee',
    'created_date', 'created_datetime',
    'updated_date', 'updated_datetime',
    'resolution', 'resolution_date', 'resolution_datetime',
    'labels', 'categoria_AQN', 'PMBOK', 'informacion_completada',
    'description',
]
col_order_view = [c for c in col_order_view if c in df_view.columns]
df_view = df_view[col_order_view]

st.dataframe(
    df_view,
    use_container_width=True,
    hide_index=True,
    column_config={
        'url_ticket': st.column_config.LinkColumn(
            label="🔗 URL Ticket",
            display_text="https://prestamype\\.atlassian\\.net/browse/(.+)",
        ),
        'issue_key':              st.column_config.TextColumn("Clave"),
        'proyecto_codigo':        st.column_config.TextColumn("Proyecto"),
        'issue_id':               st.column_config.TextColumn("ID"),
        'summary':                st.column_config.TextColumn("Resumen"),
        'status':                 st.column_config.TextColumn("Estado"),
        'issue_type':             st.column_config.TextColumn("Tipo"),
        'assignee':               st.column_config.TextColumn("Asignado a"),
        'created_date':           st.column_config.TextColumn("Fecha creación"),
        'created_datetime':       st.column_config.TextColumn("Fecha-hora creación"),
        'updated_date':           st.column_config.TextColumn("Fecha actualización"),
        'updated_datetime':       st.column_config.TextColumn("Fecha-hora actualización"),
        'resolution':             st.column_config.TextColumn("Resolución"),
        'resolution_date':        st.column_config.TextColumn("Fecha resolución"),
        'resolution_datetime':    st.column_config.TextColumn("Fecha-hora resolución"),
        'labels':                 st.column_config.TextColumn("Etiquetas"),
        'categoria_AQN':          st.column_config.TextColumn("Categoría AQN"),
        'PMBOK':                  st.column_config.TextColumn("PMBOK"),
        'informacion_completada': st.column_config.TextColumn("Info completada"),
        'description':            st.column_config.TextColumn("Descripción"),
    }
)
