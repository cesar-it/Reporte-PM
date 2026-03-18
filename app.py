# ============================================================
# EXTRACTOR JIRA - STREAMLIT
# Credenciales via st.secrets
# ============================================================

import os
import io
import time
import requests
import pandas as pd
import streamlit as st
from requests.auth import HTTPBasicAuth
from datetime import datetime
from collections import defaultdict

# ============================================================
# CONFIGURACIÓN ESTÁTICA
# ============================================================

PROJECT_KEYS = ['PM']

ASSIGNEES = [
    "Angie Tomasto", "valeria vergaray",
    "crisbel aguilar", "Miguel Carreño"
]

STATE_EN_CURSO_UX = "EN CURSO DE UX"
STATE_BACKLOG_SW       = "BACKLOG SOFTWARE | COE"
STATE_BACKLOG_SW_OLD  = "BACKLOG SOFTWARE"
STATE_EN_CURSO_SW = "EN CURSO DE SOFTWARE | COE"
STATE_ATENDIDO    = "ATENDIDO"

ALL_STATUSES = [
    "BACKLOG",
    "BACKLOG UX",
    STATE_EN_CURSO_UX,
    STATE_BACKLOG_SW,
    STATE_BACKLOG_SW_OLD,
    STATE_EN_CURSO_SW,
    STATE_ATENDIDO,
    "EN ESPERA",
    "CANCELADO",
]

# ============================================================
# CREDENCIALES DESDE st.secrets
# ============================================================

JIRA_URL  = st.secrets["JIRA_URL"]
USERNAME  = st.secrets["JIRA_USER"]
API_TOKEN = st.secrets["JIRA_TOKEN"]

# ============================================================
# CLASE: EXTRACTOR DE ISSUES
# ============================================================

class JiraExtractor:
    def __init__(self):
        self.auth    = HTTPBasicAuth(USERNAME, API_TOKEN)
        self.headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}
        self.base    = JIRA_URL.rstrip('/')

    def _post(self, url, payload):
        r = requests.post(url, json=payload, auth=self.auth, headers=self.headers)
        r.raise_for_status()
        return r.json()

    def _get(self, url, params=None):
        r = requests.get(url, params=params, auth=self.auth, headers=self.headers)
        r.raise_for_status()
        return r.json()

    def test_connection(self):
        try:
            info = self._get(f"{self.base}/rest/api/2/myself")
            return True, info.get('displayName', '')
        except Exception as e:
            return False, str(e)

    def fetch_issues(self, project_key, date_from_str, date_to_str, log_fn=None):
        url = f"{self.base}/rest/api/3/search/jql"
        assignees_jql = ', '.join(f'"{a}"' for a in ASSIGNEES)
        jql = (
            f'project = "{project_key}" '
            f'AND assignee in ({assignees_jql}) '
            f'AND created >= "{date_from_str}" '
            f'AND created <= "{date_to_str}" '
            f'ORDER BY created DESC'
        )
        all_issues = []
        next_token = None
        page = 0

        while True:
            payload = {
                "jql": jql,
                "maxResults": 50,
                "fields": [
                    "key", "summary", "status", "assignee",
                    "created", "updated", "issuetype",
                    "resolution", "resolutiondate", "description", "labels",
                    "customfield_12066", "customfield_12067", "customfield_12166"
                ]
            }
            if next_token:
                payload["nextPageToken"] = next_token

            data   = self._post(url, payload)
            issues = data.get('issues', [])
            if not issues:
                break

            page += 1
            for issue in issues:
                all_issues.append(self._parse_issue(issue, project_key))
            if log_fn:
                log_fn(f"Página {page}: {len(all_issues)} issues acumulados")

            next_token = data.get('nextPageToken')
            if not next_token:
                break
            time.sleep(0.3)

        return all_issues

    def _parse_issue(self, issue, project_key):
        f = issue.get('fields', {})

        def sg(obj, *keys):
            r = obj
            for k in keys:
                r = r.get(k) if isinstance(r, dict) else None
            return r or ''

        def fmt_dt(s):
            return (s[:10] + ' ' + s[11:19]) if s and 'T' in s else (s or '')

        def fmt_d(s):
            return s[:10] if s else ''

        def clean_desc(d):
            if not d: return ''
            if isinstance(d, dict):
                parts = []
                for block in d.get('content', []):
                    for sub in block.get('content', []):
                        if sub.get('text'): parts.append(sub['text'])
                return ' '.join(parts)[:400]
            return str(d)[:400]

        def parse_arr(fd):
            if not fd: return ''
            if isinstance(fd, list):
                return ', '.join(
                    (i.get('value', str(i)) if isinstance(i, dict) else str(i))
                    for i in fd
                )
            return str(fd)

        def parse_single(fd):
            if not fd: return ''
            return fd.get('value', '') if isinstance(fd, dict) else str(fd)

        return {
            'proyecto_codigo':        project_key,
            'issue_key':              issue.get('key', ''),
            'url_ticket':             f"https://prestamype.atlassian.net/browse/{issue.get('key', '')}",
            'issue_id':               issue.get('id', ''),
            'summary':                f.get('summary', ''),
            'description':            clean_desc(f.get('description')),
            'status':                 sg(f, 'status', 'name') or 'Sin estado',
            'issue_type':             sg(f, 'issuetype', 'name') or 'Sin tipo',
            'assignee':               sg(f, 'assignee', 'displayName') or 'Sin asignar',
            'created_date':           fmt_d(f.get('created', '')),
            'created_datetime':       fmt_dt(f.get('created', '')),
            'updated_date':           fmt_d(f.get('updated', '')),
            'updated_datetime':       fmt_dt(f.get('updated', '')),
            'resolution':             sg(f, 'resolution', 'name') or 'Sin resolver',
            'resolution_date':        fmt_d(f.get('resolutiondate', '')),
            'resolution_datetime':    fmt_dt(f.get('resolutiondate', '')),
            'labels':                 ', '.join(f.get('labels', [])),
            'categoria_AQN':          parse_arr(f.get('customfield_12066')),
            'PMBOK':                  parse_single(f.get('customfield_12067')),
            'informacion_completada': parse_single(f.get('customfield_12166')),
        }


# ============================================================
# CLASE: EXTRACTOR DE CHANGELOG
# ============================================================

class ChangelogExtractor:
    def __init__(self):
        self.auth    = HTTPBasicAuth(USERNAME, API_TOKEN)
        self.headers = {'Accept': 'application/json', 'Content-Type': 'application/json'}
        self.base    = JIRA_URL.rstrip('/')

    def fetch_changelog(self, issue_key):
        url    = f"{self.base}/rest/api/3/issue/{issue_key}/changelog"
        result = []
        start  = 0

        while True:
            try:
                r = requests.get(
                    url, params={'startAt': start, 'maxResults': 100},
                    auth=self.auth, headers=self.headers
                )
                if r.status_code != 200:
                    break
                data   = r.json()
                values = data.get('values', [])
                if not values:
                    break

                for entry in values:
                    created = entry.get('created', '')
                    author  = entry.get('author', {}).get('displayName', 'Unknown')
                    for item in entry.get('items', []):
                        if item.get('field', '').lower() == 'status':
                            result.append({
                                'issue_key':   issue_key,
                                'from_status': item.get('fromString', ''),
                                'to_status':   item.get('toString', ''),
                                'change_dt':   created[:19].replace('T', ' ') if len(created) >= 19 else created,
                                'changed_by':  author,
                            })

                if data.get('isLast', True):
                    break
                start += 100
            except Exception:
                break

        return result

    def fetch_all(self, issue_keys, progress_bar=None, log_fn=None):
        all_changes = []
        total = len(issue_keys)
        for i, key in enumerate(issue_keys, 1):
            all_changes.extend(self.fetch_changelog(key))
            if progress_bar:
                progress_bar.progress(i / total, text=f"Changelog {i}/{total} issues")
            if log_fn and (i % 20 == 0 or i == total):
                log_fn(f"Changelog {i}/{total}")
            time.sleep(0.12)
        return all_changes


# ============================================================
# LÓGICA DE NEGOCIO
# ============================================================

def compute_times(issue_keys, changelog_list):
    cl_by_issue = defaultdict(list)
    for c in changelog_list:
        cl_by_issue[c['issue_key']].append(c)
    for key in cl_by_issue:
        cl_by_issue[key].sort(key=lambda x: x['change_dt'])

    results = {}
    for key in issue_keys:
        changes  = cl_by_issue.get(key, [])
        t_ux = t_sw = ts_ux_in = ts_sw_in = None
        dt_entrada_ux = dt_salida_sw = None

        for c in changes:
            to_s   = c['to_status'].upper().strip()
            from_s = c['from_status'].upper().strip()
            try:
                dt = datetime.strptime(c['change_dt'], '%Y-%m-%d %H:%M:%S')
            except Exception:
                continue

            if to_s == STATE_EN_CURSO_UX and ts_ux_in is None:
                ts_ux_in      = dt
                dt_entrada_ux = dt.date()

            if (from_s == STATE_EN_CURSO_UX and
                    to_s in (STATE_BACKLOG_SW, STATE_BACKLOG_SW_OLD, STATE_EN_CURSO_SW) and
                    ts_ux_in is not None and t_ux is None):
                t_ux         = round((dt - ts_ux_in).total_seconds() / 3600, 2)
                dt_salida_sw = dt.date()

            if to_s == STATE_EN_CURSO_SW and ts_sw_in is None:
                ts_sw_in = dt
                if dt_salida_sw is None:
                    dt_salida_sw = dt.date()

            if (from_s == STATE_EN_CURSO_SW and
                    to_s == STATE_ATENDIDO and
                    ts_sw_in is not None and t_sw is None):
                t_sw = round((dt - ts_sw_in).total_seconds() / 3600, 2)

        results[key] = {
            'tiempo_ux_horas':  t_ux if t_ux is not None else '',
            'tiempo_sw_horas':  t_sw if t_sw is not None else '',
            'fecha_entrada_ux': str(dt_entrada_ux) if dt_entrada_ux else '',
            'fecha_salida_sw':  str(dt_salida_sw)  if dt_salida_sw  else '',
        }
    return results


def apply_filters(issues, time_map,
                  selected_statuses,
                  ux_active, ux_from, ux_to,
                  sw_active, sw_from, sw_to):
    filtered = []
    for issue in issues:
        key = issue['issue_key']
        tm  = time_map.get(key, {})

        if selected_statuses:
            if issue['status'].upper().strip() not in [s.upper().strip() for s in selected_statuses]:
                continue

        if ux_active:
            fecha_ux = tm.get('fecha_entrada_ux', '')
            if not fecha_ux:
                continue
            try:
                if not (ux_from <= datetime.strptime(fecha_ux, '%Y-%m-%d').date() <= ux_to):
                    continue
            except Exception:
                continue

        if sw_active:
            fecha_sw = tm.get('fecha_salida_sw', '')
            if not fecha_sw:
                continue
            try:
                if not (sw_from <= datetime.strptime(fecha_sw, '%Y-%m-%d').date() <= sw_to):
                    continue
            except Exception:
                continue

        filtered.append(issue)
    return filtered


def build_excel_bytes(issues, time_map, changelog_list):
    from openpyxl.styles import PatternFill, Font, Alignment

    df = pd.DataFrame(issues)
    df['tiempo_ux_horas']  = df['issue_key'].map(lambda k: time_map.get(k, {}).get('tiempo_ux_horas', ''))
    df['tiempo_sw_horas']  = df['issue_key'].map(lambda k: time_map.get(k, {}).get('tiempo_sw_horas', ''))
    df['fecha_entrada_ux'] = df['issue_key'].map(lambda k: time_map.get(k, {}).get('fecha_entrada_ux', ''))
    df['fecha_salida_sw']  = df['issue_key'].map(lambda k: time_map.get(k, {}).get('fecha_salida_sw', ''))

    col_order = [
        'proyecto_codigo', 'issue_key', 'url_ticket', 'issue_id', 'summary',
        'status', 'issue_type', 'assignee',
        'created_date', 'created_datetime',
        'updated_date', 'updated_datetime',
        'resolution', 'resolution_date', 'resolution_datetime',
        'labels', 'categoria_AQN', 'PMBOK', 'informacion_completada',
        'tiempo_ux_horas', 'fecha_entrada_ux',
        'tiempo_sw_horas', 'fecha_salida_sw',
        'description',
    ]
    col_order = [c for c in col_order if c in df.columns]
    df = df[col_order]

    keys_in_report = set(df['issue_key'].tolist())
    cl_filtered = [c for c in changelog_list if c['issue_key'] in keys_in_report]
    df_cl = pd.DataFrame(cl_filtered) if cl_filtered else pd.DataFrame(
        columns=['issue_key', 'from_status', 'to_status', 'change_dt', 'changed_by']
    )
    if not df_cl.empty:
        df_cl = df_cl.sort_values(['issue_key', 'change_dt']).reset_index(drop=True)
    df_cl.columns = ['Issue Key', 'De Estado', 'A Estado', 'Fecha y Hora', 'Modificado por']

    def style_header(ws, columns, special_cols, sp_fill, sp_font_color):
        for i, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=i)
            is_sp          = col in special_cols
            cell.fill      = PatternFill("solid", fgColor=sp_fill if is_sp else "1a73e8")
            cell.font      = Font(bold=True, color=sp_font_color if is_sp else "FFFFFF", size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    col_w_issues = {
        'proyecto_codigo': 16, 'issue_key': 14, 'issue_id': 12, 'url_ticket': 45, 'summary': 40,
        'status': 20, 'issue_type': 16, 'assignee': 22,
        'created_date': 14, 'created_datetime': 20,
        'updated_date': 14, 'updated_datetime': 20,
        'resolution': 16, 'resolution_date': 14, 'resolution_datetime': 20,
        'labels': 20, 'categoria_AQN': 22, 'PMBOK': 18,
        'informacion_completada': 24,
        'tiempo_ux_horas': 18, 'fecha_entrada_ux': 18,
        'tiempo_sw_horas': 18, 'fecha_salida_sw': 18,
        'description': 50,
    }
    col_w_cl = {
        'Issue Key': 14, 'De Estado': 30, 'A Estado': 30,
        'Fecha y Hora': 22, 'Modificado por': 24,
    }

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Hoja 1: Issues
        df.to_excel(writer, index=False, sheet_name='Issues')
        ws_i = writer.sheets['Issues']
        style_header(ws_i, df.columns,
                     {'tiempo_ux_horas', 'fecha_entrada_ux', 'tiempo_sw_horas', 'fecha_salida_sw'},
                     'FFF2CC', '7F6000')
        for i, col in enumerate(df.columns, 1):
            ws_i.column_dimensions[ws_i.cell(row=1, column=i).column_letter].width = col_w_issues.get(col, 18)
        ws_i.freeze_panes = 'A2'

        # Hoja 2: Changelog
        df_cl.to_excel(writer, index=False, sheet_name='Changelog')
        ws_c = writer.sheets['Changelog']
        style_header(ws_c, df_cl.columns,
                     {'De Estado', 'A Estado'}, 'E8F0FE', '1a3e8a')
        for i, col in enumerate(df_cl.columns, 1):
            ws_c.column_dimensions[ws_c.cell(row=1, column=i).column_letter].width = col_w_cl.get(col, 20)
        ws_c.freeze_panes = 'A2'

    output.seek(0)
    return output.read()


# ============================================================
# UI STREAMLIT
# ============================================================

st.set_page_config(page_title="Extractor JIRA", page_icon="📊", layout="wide")
st.title("📊 Extractor JIRA — Reporte TD")

# ── Sidebar ─────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Configuración")

    st.subheader("📅 Creación del ticket")
    col_a, col_b = st.columns(2)
    with col_a:
        date_from = st.date_input("Desde", value=datetime(2025, 12, 1).date())
    with col_b:
        date_to   = st.date_input("Hasta", value=datetime(2025, 12, 31).date())

    st.subheader("🔖 Estado del ticket")
    selected_statuses = st.multiselect(
        "Estados (vacío = todos)",
        options=ALL_STATUSES,
        default=[]
    )

    st.subheader("🎨 Transición → EN CURSO UX")
    ux_active = st.checkbox("Activar filtro UX", value=False)
    ux_from   = st.date_input("UX Desde", value=datetime(2025, 12, 1).date(),
                               disabled=not ux_active, key="ux_from")
    ux_to     = st.date_input("UX Hasta", value=datetime(2025, 12, 31).date(),
                               disabled=not ux_active, key="ux_to")

    st.subheader("💻 Transición → BACKLOG / EN CURSO SW")
    sw_active = st.checkbox("Activar filtro SW", value=False)
    sw_from   = st.date_input("SW Desde", value=datetime(2025, 12, 1).date(),
                               disabled=not sw_active, key="sw_from")
    sw_to     = st.date_input("SW Hasta", value=datetime(2025, 12, 31).date(),
                               disabled=not sw_active, key="sw_to")

    st.subheader("💾 Nombre del archivo")
    filename_input = st.text_input("Nombre (sin .xlsx)", value="reporte_jira_TD")

    run_btn = st.button("🚀 Ejecutar extracción", type="primary", use_container_width=True)

# ── Panel principal ──────────────────────────────────────────
if not run_btn:
    st.info("Configura los filtros en el panel izquierdo y presiona **Ejecutar extracción**.")
    st.stop()

# Validación de fechas
if date_from > date_to:
    st.error("La fecha de inicio no puede ser mayor a la fecha fin.")
    st.stop()

# Conexión
extractor = JiraExtractor()
ok, display_name = extractor.test_connection()
if not ok:
    st.error(f"Error de conexión con JIRA: {display_name}")
    st.stop()
st.success(f"✅ Conectado como: {display_name}")

ch_extractor = ChangelogExtractor()
all_issues   = []

# Paso 1: Issues
with st.status("Extrayendo issues...", expanded=True) as status_issues:
    for pk in PROJECT_KEYS:
        st.write(f"Proyecto: **{pk}**")
        issues = extractor.fetch_issues(
            pk,
            date_from.strftime('%Y-%m-%d'),
            date_to.strftime('%Y-%m-%d'),
            log_fn=lambda m: st.write(m)
        )
        all_issues.extend(issues)
    status_issues.update(
        label=f"✅ Issues extraídos: {len(all_issues)}",
        state="complete"
    )

if not all_issues:
    st.warning("No se encontraron issues en el rango de fechas seleccionado.")
    st.stop()

# Paso 2: Changelog
issue_keys = [i['issue_key'] for i in all_issues]
with st.status("Extrayendo changelog...", expanded=True) as status_cl:
    prog_bar = st.progress(0, text="Iniciando...")
    changelog = ch_extractor.fetch_all(issue_keys, progress_bar=prog_bar)
    status_cl.update(
        label=f"✅ Changelog extraído: {len(changelog)} cambios de estado",
        state="complete"
    )

# Paso 3: Tiempos + filtros
with st.spinner("Calculando tiempos y aplicando filtros..."):
    time_map        = compute_times(issue_keys, changelog)
    filtered_issues = apply_filters(
        all_issues, time_map,
        selected_statuses,
        ux_active, ux_from, ux_to,
        sw_active, sw_from, sw_to,
    )

if not filtered_issues:
    st.warning("Ningún issue pasó los filtros aplicados. Ajusta los criterios en el sidebar.")
    st.stop()

# Métricas
ux_count = sum(1 for v in time_map.values() if v['tiempo_ux_horas'] != '')
sw_count = sum(1 for v in time_map.values() if v['tiempo_sw_horas'] != '')
cl_count = len([c for c in changelog if c['issue_key'] in {i['issue_key'] for i in filtered_issues}])

m1, m2, m3, m4 = st.columns(4)
m1.metric("Issues en reporte",    len(filtered_issues))
m2.metric("Con Tiempo UX (hrs)",  ux_count)
m3.metric("Con Tiempo SW (hrs)",  sw_count)
m4.metric("Registros changelog",  cl_count)

# Paso 4: Generar Excel
with st.spinner("Generando archivo Excel..."):
    excel_bytes = build_excel_bytes(filtered_issues, time_map, changelog)

safe_name = "".join(c for c in filename_input.strip() if c.isalnum() or c in ('_', '-')).strip() or 'jira_report'
excel_filename = f"{safe_name}.xlsx"

st.download_button(
    label="⬇️ Descargar reporte Excel",
    data=excel_bytes,
    file_name=excel_filename,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)

# Vista previa
with st.expander("👁️ Vista previa de issues (primeras 50 filas)"):
    df_preview = pd.DataFrame(filtered_issues)[
        ['issue_key', 'summary', 'status', 'assignee', 'created_date']
    ].copy().head(50)
    df_preview['tiempo_ux_horas'] = df_preview['issue_key'].map(
        lambda k: time_map.get(k, {}).get('tiempo_ux_horas', ''))
    df_preview['tiempo_sw_horas'] = df_preview['issue_key'].map(
        lambda k: time_map.get(k, {}).get('tiempo_sw_horas', ''))
    st.dataframe(df_preview, use_container_width=True)
